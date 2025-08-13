# search_logic.py
import re
from openpyxl import load_workbook
from openpyxl.styles.colors import Color
from config import RELATIVE_OFFSETS
from utils import parse_amount

def clean_text(text):
    """정규표현식을 사용하여 텍스트를 빠르게 정리합니다."""
    if not isinstance(text, str):
        return text
    cleaned_text = re.sub(r'[\s\x00-\x1F\x7F]+', ' ', text)
    return cleaned_text.strip()

def get_status_from_color(color_obj) -> str:
    """셀의 색상 객체를 분석하여 데이터 상태 텍스트("최신" 등)로 변환합니다."""
    if not isinstance(color_obj, Color): return "미지정"
    if color_obj.type == 'theme':
        if color_obj.theme == 6: return "최신"
        if color_obj.theme == 3: return "1년 경과"
        if color_obj.theme in [0, 1]: return "1년 이상 경과"
    elif color_obj.type == 'rgb':
        hex_color = color_obj.rgb.upper() if color_obj.rgb else "00000000"
        if hex_color == "FFE2EFDA": return "최신"
        if hex_color == "FFDDEBF7": return "1년 경과"
        if hex_color in ["FFFFFFFF", "00000000", "FFFDEDEC"]: return "1년 이상 경과"
    return "미지정"

# [find_and_filter_companies 함수를 이 코드로 통째로 교체하세요]
def find_and_filter_companies(file_path, filters):
    all_companies = []
    
    try:
        workbook = load_workbook(filename=file_path, data_only=False)
    except Exception as e:
        return [{"오류": f"파일 열기 오류: {e}"}]

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        for r_idx, row_cells in enumerate(sheet.iter_rows(max_row=max_row, max_col=max_col)):
            excel_row_num = r_idx + 1
            
            first_cell_value = row_cells[0].value
            if isinstance(first_cell_value, str) and "회사명" in first_cell_value.strip():
                
                for c_idx, company_header_cell in enumerate(row_cells[1:], start=1):
                    excel_col_num = company_header_cell.column
                    company_name = company_header_cell.value
                    
                    if not isinstance(company_name, str) or not company_name.strip():
                        continue

                    cleaned_company_name = clean_text(company_name)
                    company_data = {"검색된 회사": cleaned_company_name}
                    
                    for item, offset in RELATIVE_OFFSETS.items():
                        target_row = excel_row_num + offset
                        if target_row <= max_row:
                            cell = sheet.cell(row=target_row, column=excel_col_num)
                            value = cell.value
                            if item in ["부채비율", "유동비율"]:
                                if isinstance(value, (int, float)): processed_value = value * 100
                                elif isinstance(value, str):
                                    try: processed_value = float(value.replace('%', '').strip())
                                    except (ValueError, TypeError): processed_value = clean_text(value)
                                else: processed_value = value


                            elif item == "신용평가":
                                if isinstance(value, str):
                                    # 1. 양쪽 끝의 불필요한 공백을 모두 제거합니다.
                                    cleaned_value = value.strip()
                                    # 2. 중간에 있을지 모르는 여러 개의 공백/특수문자를 단일 공백으로 바꿉니다.
                                    normalized_value = " ".join(cleaned_value.split())
                                    # 3. 완전히 정리된 문자열에서 첫 공백을 줄바꿈으로 변경합니다.
                                    processed_value = normalized_value.replace(' ', '\n', 1)
                                else:
                                    processed_value = value
                            else:
                                processed_value = clean_text(value) if isinstance(value, str) else value
                            company_data[item] = processed_value if processed_value is not None else ""
                        else:
                            company_data[item] = "N/A"
                    
                    company_statuses = {}
                    for item, offset in RELATIVE_OFFSETS.items():
                        target_row = excel_row_num + offset
                        if target_row <= max_row:
                            cell = sheet.cell(row=target_row, column=excel_col_num)
                            company_statuses[item] = get_status_from_color(cell.fill.fgColor if cell.fill else None)
                        else:
                            company_statuses[item] = "범위 초과"
                    company_data["데이터상태"] = company_statuses
                    
                    all_companies.append(company_data)

    if not all_companies:
        return [{"오류": "엑셀 파일에서 업체 정보를 찾을 수 없습니다."}]
    
    # --- 필터링 로직 ---
    filtered_results = all_companies
    if filters.get('name'):
        search_name = filters['name'].lower()
        filtered_results = [comp for comp in filtered_results if search_name in str(comp.get("검색된 회사", "")).lower()]
    
    # [핵심] 담당자 필터링 로직 추가
    if filters.get('manager'):
        search_manager = filters['manager'].lower()
        # '비고' 필드에 담당자 이름이 포함되어 있는지 확인 (대소문자 구분 없음)
        filtered_results = [comp for comp in filtered_results if search_manager in str(comp.get("비고", "")).lower()]

    if filters.get('region') and filters['region'] != "전체":
        search_region = filters['region'].lower()
        filtered_results = [comp for comp in filtered_results if search_region in str(comp.get("지역", "")).lower()]
    
    for key, field_name in [('sipyung', '시평'), ('perf_3y', '3년 실적'), ('perf_5y', '5년 실적')]:
        min_val, max_val = filters.get(f'min_{key}'), filters.get(f'max_{key}')
        if min_val is not None:
            filtered_results = [comp for comp in filtered_results if (val := parse_amount(str(comp.get(field_name)))) is not None and val >= min_val]
        if max_val is not None:
            filtered_results = [comp for comp in filtered_results if (val := parse_amount(str(comp.get(field_name)))) is not None and val <= max_val]

    if not filtered_results:
        return [{"오류": "주어진 조건에 맞는 업체를 찾을 수 없습니다."}]
        
    return filtered_results