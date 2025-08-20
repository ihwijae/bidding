# ui_pyside/result_management_dialog.py

import os
import json
import copy
import sys
import re

from openpyxl.styles import PatternFill, Alignment

import calculation_logic
import utils
from datetime import datetime
from PySide6.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QGroupBox,
                               QPushButton, QTableWidget, QTableWidgetItem,
                               QHeaderView, QSplitter, QLabel, QMessageBox,
                               QInputDialog, QGridLayout, QFileDialog)
from PySide6.QtCore import Qt, Signal
from .load_consortium_popup import LoadConsortiumPopup
from PySide6.QtWidgets import QFrame # QFrame 추가
from .text_display_popup import TextDisplayPopup
from consortium_manager import ConsortiumManagerDialog
from openpyxl import load_workbook


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)



class ResultManagementDialog(QDialog):
    # 변경된 결과 목록을 메인 윈도우로 다시 전달하기 위한 시그널
    results_updated = Signal(list)

    def __init__(self, result_widgets, controller, region_limit, parent=None):
        super().__init__(parent)
        self.result_widgets = result_widgets  # 메인 윈도우의 결과 목록을 복사해 옴
        self.controller = controller  # 메인 윈도우의 기능(엑셀 저장 등)을 호출하기 위함
        self.region_limit = region_limit


        self.setWindowTitle("협정 결과 관리")
        self.setMinimumSize(1200, 600)



        # UI 설정
        self.setup_ui()
        # 데이터 채우기
        self.populate_consortium_list()

    def setup_ui(self):
        main_layout = QVBoxLayout(self)

        # 1. 상단 툴바 (이전과 동일)
        toolbar_layout = QHBoxLayout()
        self.save_button = QPushButton("💾 현재 목록 저장")
        self.generate_messages_button = QPushButton("✉️ 협정 문자 생성")
        self.excel_export_button = QPushButton("📊 엑셀로 내보내기")
        toolbar_layout.addWidget(self.save_button)
        toolbar_layout.addWidget(self.generate_messages_button)
        toolbar_layout.addStretch(1)
        toolbar_layout.addWidget(self.excel_export_button)
        main_layout.addLayout(toolbar_layout)

        # 2. 메인 컨텐츠 (좌/우 분할)
        splitter = QSplitter(Qt.Orientation.Horizontal)

        # 2-1. 좌측 패널 (컨소시엄 목록)
        left_panel = QGroupBox("컨소시엄 목록")
        left_layout = QVBoxLayout(left_panel)
        self.consortium_list_table = QTableWidget()
        self.consortium_list_table.setColumnCount(4)
        self.consortium_list_table.setHorizontalHeaderLabels(["No.", "대표사", "구성사 수", "종합점수"])
        self.consortium_list_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.consortium_list_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.consortium_list_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)

        # ▼▼▼▼▼ [핵심] '협정 상세 수정' 버튼을 여기에 추가합니다 ▼▼▼▼▼
        list_action_layout = QHBoxLayout()
        self.review_button = QPushButton("🔍 상세 검토")
        self.detailed_edit_button = QPushButton("📝 협정 상세 수정")  # <-- 새로 추가!
        self.move_up_button = QPushButton("▲ 위로")
        self.move_down_button = QPushButton("▼ 아래로")
        self.duplicate_button = QPushButton("📄 선택 복제")
        self.delete_button = QPushButton("❌ 선택 삭제")
        list_action_layout.addStretch(1)
        list_action_layout.addWidget(self.review_button)
        list_action_layout.addWidget(self.detailed_edit_button)  # <-- 레이아웃에 추가!
        list_action_layout.addWidget(self.move_up_button)
        list_action_layout.addWidget(self.move_down_button)
        list_action_layout.addWidget(self.duplicate_button)
        list_action_layout.addWidget(self.delete_button)
        # ▲▲▲▲▲ 여기까지 ▲▲▲▲▲

        left_layout.addWidget(self.consortium_list_table)
        left_layout.addLayout(list_action_layout)

        # 2-2. 우측 패널 (상세 정보) (이전과 동일)
        right_panel = QGroupBox("선택한 컨소시엄 상세정보")
        right_layout = QVBoxLayout(right_panel)
        self.detail_title_label = QLabel("← 왼쪽 목록에서 협정을 선택하세요.")
        self.detail_table = QTableWidget()
        self.detail_table.setColumnCount(5)
        self.detail_table.setHorizontalHeaderLabels(["구분", "업체명", "지분율(%)", "경영점수", "5년실적"])
        self.detail_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.detail_table.setFixedHeight(150)

        # ▼▼▼ [핵심 추가] 점수 요약 그룹박스 ▼▼▼
        score_summary_group = QGroupBox("점수 요약")
        score_layout = QGridLayout(score_summary_group)

        # 각 점수를 표시할 QLabel들을 생성합니다.
        self.final_biz_score_label = QLabel("N/A")
        self.performance_ratio_label = QLabel("N/A")
        self.final_perf_score_label = QLabel("N/A")
        self.total_score_label = QLabel("N/A")

        # 그리드 레이아웃에 제목과 값 QLabel을 배치합니다.
        score_layout.addWidget(QLabel("<b>경영점수 총점:</b>"), 0, 0)
        score_layout.addWidget(self.final_biz_score_label, 0, 1)
        score_layout.addWidget(QLabel("<b>실적 비율:</b>"), 1, 0)
        score_layout.addWidget(self.performance_ratio_label, 1, 1)
        score_layout.addWidget(QLabel("<b>실적 점수:</b>"), 2, 0)
        score_layout.addWidget(self.final_perf_score_label, 2, 1)
        score_layout.addWidget(QLabel("<b>예상 종합 평점:</b>"), 3, 0)
        score_layout.addWidget(self.total_score_label, 3, 1)

        right_layout.addWidget(self.detail_title_label)
        right_layout.addWidget(self.detail_table)
        right_layout.addWidget(score_summary_group)  # 우측 패널에 점수 요약 그룹 추가
        right_layout.addStretch(1)  # 하단 공간 확보


        splitter.addWidget(left_panel)
        splitter.addWidget(right_panel)
        splitter.setSizes([500, 680])
        main_layout.addWidget(splitter)

        # 3. 하단 닫기 버튼
        self.close_button = QPushButton("저장하고 닫기")
        main_layout.addWidget(self.close_button, 0, Qt.AlignmentFlag.AlignRight)

        # 4. 시그널 연결
        self.consortium_list_table.itemSelectionChanged.connect(self.update_detail_view)
        self.close_button.clicked.connect(self.accept)
        self.delete_button.clicked.connect(self.delete_selected_consortium)
        self.review_button.clicked.connect(self.open_review_for_selected)
        self.detailed_edit_button.clicked.connect(self._open_consortium_editor)  # <-- 새 버튼 연결 추가!
        self.duplicate_button.clicked.connect(self.duplicate_selected_consortium)
        self.move_up_button.clicked.connect(self.move_consortium_up)
        self.move_down_button.clicked.connect(self.move_consortium_down)
        self.excel_export_button.clicked.connect(self.generate_excel_report)
        self.generate_messages_button.clicked.connect(self.generate_consortium_messages)
        self.save_button.clicked.connect(self.save_consortiums_list)

    def populate_consortium_list(self):
        """좌측 목록 테이블을 현재 데이터로 채웁니다."""
        self.consortium_list_table.setRowCount(len(self.result_widgets))
        for i, widget in enumerate(self.result_widgets):
            data = widget.result_data
            details = data.get("company_details", [])
            lead_company = details[0].get("name", "N/A") if details else "N/A"
            num_members = len(details)
            total_score = data.get("expected_score", 0)

            self.consortium_list_table.setItem(i, 0, QTableWidgetItem(str(i + 1)))
            self.consortium_list_table.setItem(i, 1, QTableWidgetItem(lead_company))
            self.consortium_list_table.setItem(i, 2, QTableWidgetItem(f"{num_members} 개사"))
            self.consortium_list_table.setItem(i, 3, QTableWidgetItem(f"{total_score:.4f}"))
        self.consortium_list_table.resizeColumnsToContents()

    def update_detail_view(self):
        """선택된 컨소시엄의 상세 정보와 '점수 요약'을 우측에 표시합니다."""
        selected_rows = self.consortium_list_table.selectionModel().selectedRows()
        if not selected_rows:
            self.detail_title_label.setText("← 왼쪽 목록에서 협정을 선택하세요.")
            self.detail_table.setRowCount(0)
            # 선택이 없으면 점수 요약도 초기화
            self.final_biz_score_label.setText("N/A")
            self.performance_ratio_label.setText("N/A")
            self.final_perf_score_label.setText("N/A")
            self.total_score_label.setText("N/A")
            return

        selected_row = selected_rows[0].row()
        data = self.result_widgets[selected_row].result_data
        details = data.get("company_details", [])

        self.detail_title_label.setText(f"<b>No. {selected_row + 1} 상세정보</b> | {data.get('gongo_title', '')}")
        self.detail_table.setRowCount(len(details))

        # 1. 업체 목록 테이블을 채웁니다.
        for i, comp in enumerate(details):
            share_percent = comp.get('share', 0) * 100.0
            self.detail_table.setItem(i, 0, QTableWidgetItem(comp.get('role', '')))
            self.detail_table.setItem(i, 1, QTableWidgetItem(comp.get('name', '')))
            self.detail_table.setItem(i, 2, QTableWidgetItem(f"{share_percent:.2f}%"))
            self.detail_table.setItem(i, 3,
                                      QTableWidgetItem(f"{comp.get('business_score_details', {}).get('total', 0):.4f}"))
            self.detail_table.setItem(i, 4, QTableWidgetItem(f"{comp.get('performance_5y', 0):,}"))

        # ▼▼▼ [핵심 수정] for 반복문이 끝난 후, 딱 한 번만 점수 요약을 업데이트합니다. ▼▼▼
        self.final_biz_score_label.setText(f"<b>{data.get('final_business_score', 0):.4f}</b> 점")
        self.performance_ratio_label.setText(f"{data.get('performance_ratio', 0):.2f} %")
        self.final_perf_score_label.setText(f"<b>{data.get('final_performance_score', 0):.4f}</b> 점")
        self.total_score_label.setText(
            f"<span style='color:blue; font-weight:bold;'>{data.get('expected_score', 0):.4f}</span> 점")
        # ▲▲▲▲▲ [핵심 수정] 여기까지 ▲▲▲▲▲

        self.detail_table.resizeColumnsToContents()

    def delete_selected_consortium(self):
        selected_rows = self.consortium_list_table.selectionModel().selectedRows()
        if not selected_rows: return

        reply = QMessageBox.question(self, "삭제 확인", "선택한 협정 결과를 목록에서 삭제하시겠습니까?")
        if reply == QMessageBox.StandardButton.Yes:
            row_to_delete = selected_rows[0].row()
            del self.result_widgets[row_to_delete]
            self.populate_consortium_list()
            self.update_detail_view()

    def duplicate_selected_consortium(self):
        selected_rows = self.consortium_list_table.selectionModel().selectedRows()
        if not selected_rows: return

        row_to_copy = selected_rows[0].row()
        widget_to_copy = self.result_widgets[row_to_copy]

        # 깊은 복사를 통해 완전한 사본 생성
        new_widget = copy.deepcopy(widget_to_copy)

        self.result_widgets.insert(row_to_copy + 1, new_widget)
        self.populate_consortium_list()

    def move_consortium_up(self):
        selected_rows = self.consortium_list_table.selectionModel().selectedRows()
        if not selected_rows: return

        row = selected_rows[0].row()
        if row > 0:
            self.result_widgets.insert(row - 1, self.result_widgets.pop(row))
            self.populate_consortium_list()
            self.consortium_list_table.selectRow(row - 1)

    def move_consortium_down(self):
        selected_rows = self.consortium_list_table.selectionModel().selectedRows()
        if not selected_rows: return

        row = selected_rows[0].row()
        if row < len(self.result_widgets) - 1:
            self.result_widgets.insert(row + 1, self.result_widgets.pop(row))
            self.populate_consortium_list()
            self.consortium_list_table.selectRow(row + 1)

    def accept(self):
        # ▼▼▼ [진단 1] 이 print 문을 추가해주세요 ▼▼▼
        print(">>> [진단 1] accept() 실행됨. 곧 results_updated 신호를 보냅니다.")
        # ▲▲▲▲▲ 여기까지 ▲▲▲▲▲

        # 창이 닫힐 때, 변경된 목록을 메인 윈도우에 알림
        self.results_updated.emit(self.result_widgets)
        super().accept()

    def open_review_for_selected(self):
        """선택한 협정의 상세 검토창(review_dialog)을 엽니다."""
        selected_rows = self.consortium_list_table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "선택 오류", "먼저 목록에서 검토할 협정을 선택하세요.")
            return

        selected_row = selected_rows[0].row()
        result_data = self.result_widgets[selected_row].result_data

        # 기존의 ReviewDialogPyside 클래스를 재사용
        # .review_dialog import 문이 파일 상단에 필요할 수 있습니다.
        from .review_dialog import ReviewDialogPyside
        dialog = ReviewDialogPyside(result_data, self)
        dialog.exec()



    def save_consortiums_list(self):
        """현재 목록에 있는 협정들을 '이름.json' 파일 하나로 저장합니다."""
        if not self.result_widgets:
            QMessageBox.warning(self, "알림", "저장할 협정 결과가 없습니다.")
            return

        save_name, ok = QInputDialog.getText(self, "협정 파일 저장", "저장할 파일 이름을 입력하세요:")
        if not ok or not save_name.strip():
            return

        safe_filename = "".join(c for c in save_name if c not in r'<>:"/\|?*') + ".json"

        current_mode = self.controller.mode
        data_folder = os.path.join("saved_data", current_mode)
        os.makedirs(data_folder, exist_ok=True)
        file_path = os.path.join(data_folder, safe_filename)

        if os.path.exists(file_path):
            reply = QMessageBox.question(self, "덮어쓰기 확인", f"'{safe_filename}' 파일이 이미 있습니다. 덮어쓰시겠습니까?")
            if reply == QMessageBox.StandardButton.No:
                return

        region_limit = self.controller.region_limit_combo.currentText()
        project_type = self.controller.gongo_field_combo.currentText()

        data_to_save = {
            "saved_name": save_name,
            "saved_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "region_limit": region_limit,
            "project_type": project_type,
            "consortiums": [widget.result_data for widget in self.result_widgets]
        }

        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data_to_save, f, indent=4, ensure_ascii=False)
            QMessageBox.information(self, "저장 완료", f"'{safe_filename}' 이름으로 협정을 저장했습니다.")
        except Exception as e:
            QMessageBox.critical(self, "저장 실패", f"파일 저장 중 오류가 발생했습니다:\n{e}")


    def load_consortiums_list(self):
        """저장된 협정 목록(result_data)을 불러와 목록에 추가합니다."""
        popup = LoadConsortiumPopup(self.controller.mode, self)
        if not popup.exec():
            return

        selected_data = popup.get_selected_data()
        if not selected_data:
            return

        # ▼▼▼▼▼ [핵심 수정] 저장된 result_data를 바로 위젯으로 만들어 추가 ▼▼▼▼▼
        new_widgets = []
        # 키 이름을 consortiums -> saved_results로 변경
        for result_data in selected_data.get("saved_results", []):
            widget = QFrame()
            widget.result_data = result_data
            new_widgets.append(widget)

        self.result_widgets.extend(new_widgets)
        self.populate_consortium_list()
        self.update_summary_display()  # 요약 창도 업데이트
        QMessageBox.information(self, "불러오기 완료", f"'{selected_data.get('saved_name')}' 협정을 불러왔습니다.")
        # ▲▲▲▲▲ [핵심 수정] 여기까지 ▲▲▲▲▲

    def generate_consortium_messages(self):
        """요청된 최종 양식에 맞춰 협정 안내 문자를 생성합니다."""
        if not self.result_widgets:
            QMessageBox.warning(self, "알림", "먼저 '결과 표 추가' 버튼으로 계산 결과를 추가해주세요.")
            return

        all_messages_parts = []
        for result_widget in self.result_widgets:
            if not hasattr(result_widget, 'result_data'): continue

            result_data = result_widget.result_data
            gongo_no = result_data.get('gongo_no', 'N/A')
            gongo_title = result_data.get('gongo_title', 'N/A')

            message_parts = [f"{gongo_no} {gongo_title}", ""]

            details = result_data.get("company_details", [])

            for index, comp_detail in enumerate(details):
                name = comp_detail.get('name', 'N/A')
                share_decimal = comp_detail.get('share', 0)
                share_percent = share_decimal * 100.0

                line = f"{name} {'%g' % share_percent}%"
                if index < len(details) - 1:
                    line += ","

                role = comp_detail.get('role', '구성사')
                if role != "대표사":
                    biz_no = comp_detail.get('data', {}).get('사업자번호', '번호없음')
                    line += f" [{biz_no}]"
                message_parts.append(line)

            message_parts.append("")

            # [수정] 단독/협정에 따라 다른 마무리 문구를 각 메시지 블록에 추가
            if len(details) == 1:
                message_parts.append("입찰참여 부탁드립니다")
            else:
                message_parts.append("협정 부탁드립니다")

            all_messages_parts.append("\n".join(message_parts))

        # [수정] 여러 메시지를 구분선으로 연결
        final_text = "\n\n---------------------\n\n".join(all_messages_parts)

        popup = TextDisplayPopup("협정 안내 문자 (전체 복사)", final_text, self)
        popup.exec()

    def generate_excel_report(self):
        """사용자가 제공한 최종 보고서 양식(시공실적 포함)에 맞춰 엑셀 파일을 생성합니다."""
        if not self.result_widgets:
            QMessageBox.warning(self, "알림", "먼저 '결과 표 추가' 버튼으로 내보낼 결과를 추가해주세요.")
            return

        # 1. 파일 저장 경로 설정
        safe_title = "".join(c for c in self.controller.gongo_title_entry.text() if c not in r'<>:"/\|?*')
        default_filename = f"{safe_title}.xlsx"
        save_path, _ = QFileDialog.getSaveFileName(self, "엑셀 보고서 저장", default_filename, "Excel Files (*.xlsx)")
        if not save_path:
            return

        try:
            # 2. 템플릿 파일 불러오기
            template_path = resource_path("haeng_template.xlsx")
            wb = load_workbook(template_path)
            ws = wb.active

            # 3. 상단 고정 정보 채우기
            # [수정] self. -> self.controller. 로 변경
            ws['D2'] = utils.parse_amount(self.controller.estimation_price_entry.text())
            ws['M1'] = f"{self.controller.gongo_no_entry.text()} {self.controller.gongo_title_entry.text()}"
            if self.controller.bid_opening_date and self.controller.bid_opening_date.isValid():
                ws['P2'] = self.controller.bid_opening_date.toString("yyyy-MM-dd HH:mm")

                # 4. 데이터 채우기
                yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                # [수정] self. -> self.controller. 로 변경
                region_limit = self.controller.region_limit_combo.currentText()
                wrap_alignment = Alignment(vertical='center', wrap_text=True)


            # 목록에 있는 모든 컨소시엄 결과에 대해 반복 (5행부터 시작)
            for index, result_widget in enumerate(self.result_widgets):
                current_row = 5 + index
                result_data = result_widget.result_data
                details = result_data.get("company_details", [])

                # 업체별 상세정보 기록
                for comp_detail in details:
                    role = comp_detail.get('role')

                    # 1. 업체명에서 법인 형태 제거 (기존과 동일)
                    original_name = comp_detail.get('name', '')
                    company_name = re.sub(r'\s*㈜\s*|\s*\((주|유|합|재)\)\s*|\s*(주|유|합|재)식회사\s*', '', original_name).strip()

                    # ▼▼▼▼▼ [핵심 추가] 비고란에서 담당자 이름 추출 ▼▼▼▼▼
                    remarks = comp_detail.get('data', {}).get('비고', '')
                    manager_name = None
                    if remarks:
                        # '김OO', '김OO팀장' 등 2~4글자의 한글 이름을 찾는 정규표현식
                        match = re.search(r'([가-힣]{2,4})(님|팀장|실장|부장|과장|대리|주임|사원)?', remarks)
                        if match:
                            manager_name = match.group(1) # '김장섭' 부분만 추출

                        # [디버깅용 코드 추가]
                        print(f"회사: {company_name}, 비고: '{remarks}', 추출된 담당자: {manager_name}")

                    # 최종적으로 셀에 들어갈 텍스트 조합
                    final_cell_text = company_name
                    if manager_name:
                        final_cell_text += f"\n{manager_name}" # 줄바꿈 문자로 이름 추가

                    company_region = comp_detail.get('data', {}).get('지역', '')

                    if role == "대표사":
                        cell = ws.cell(current_row, 3, value=final_cell_text)
                        cell.alignment = wrap_alignment# C열
                    elif role.startswith("구성사"):
                        try:
                            col_offset = 3 + int(role.split(' ')[1])
                            cell = ws.cell(current_row, col_offset, value=final_cell_text)
                            cell.alignment = wrap_alignment
                        except:
                            continue

                    if region_limit != "전체" and region_limit in company_region:
                        cell.fill = yellow_fill

                    # I,J,K... : 지분율
                    share = comp_detail.get('share', 0)

                    # ▼▼▼▼▼ [디버깅] 엑셀에 쓰기 직전의 값을 확인합니다 ▼▼▼▼▼
                    print(f"[디버깅] 엑셀에 쓸 지분율 값: {share} (타입: {type(share)})")
                    # ▲▲▲▲▲ 여기까지 추가 ▲▲▲▲▲

                    if role == "대표사":
                        # [수정] 숫자 값을 그대로 셀에 쓰고, 셀 서식은 '백분율'로 지정
                        ws.cell(current_row, 9, value=share).number_format = '0.00%'
                    elif role.startswith("구성사"):
                        try:
                            col_offset = 9 + int(role.split(' ')[1])
                            ws.cell(current_row, col_offset, value=share).number_format = '0.00%'
                        except:
                            continue

                    # P,Q,R... : 경영상태 점수
                    biz_details = comp_detail.get('business_score_details', {})
                    biz_score = biz_details.get('total', 0)
                    if role == "대표사":
                        ws.cell(current_row, 16, value=biz_score)  # P열
                    elif role.startswith("구성사"):
                        try:
                            col_offset = 16 + int(role.split(' ')[1])
                            ws.cell(current_row, col_offset, value=biz_score)
                        except:
                            continue

                    # ▼▼▼▼▼ [추가] W,X,Y... : 5년 실적 ▼▼▼▼▼
                    performance_5y = comp_detail.get('performance_5y', 0)
                    if role == "대표사":
                        ws.cell(current_row, 23, value=performance_5y).number_format = '#,##0'  # W열
                    elif role.startswith("구성사"):
                        try:
                            col_offset = 23 + int(role.split(' ')[1])  # X, Y, Z...열
                            ws.cell(current_row, col_offset, value=performance_5y).number_format = '#,##0'
                        except:
                            continue
                    # ▲▲▲▲▲ [추가] 여기까지 ▲▲▲▲▲

            # 5. 파일 저장
            wb.save(save_path)
            QMessageBox.information(self, "성공", f"엑셀 보고서가 성공적으로 저장되었습니다.\n경로: {save_path}")

        except FileNotFoundError:
            QMessageBox.critical(self, "템플릿 파일 오류",
                                 f"템플릿 파일('haeng_template.xlsx')을 찾을 수 없습니다.\n프로젝트 폴더에 파일이 있는지 확인해주세요.")
        except Exception as e:
            QMessageBox.critical(self, "오류", f"파일 저장 중 오류가 발생했습니다: {e}")

    def _open_consortium_editor(self):
        """'상세 수정' 시점에 점수 계산에 필요한 모든 정보를 수집하여 전달합니다."""

        if not self.result_widgets:
            QMessageBox.warning(self, "알림", "수정할 협정 목록이 없습니다.")
            return

            # ▼▼▼ [핵심 수정] validate_inputs() 호출 대신, 필요한 정보만 직접 수집합니다 ▼▼▼
        try:
            # 컨트롤러(ConsortiumViewHaeng)의 UI 요소에서 직접 값을 읽어옵니다.
            announcement_date = self.controller.announcement_date_edit.date().toPython()
            rule_info = (self.controller.mode, self.controller.rule_combo.currentText())

            estimation_price = utils.parse_amount(self.controller.estimation_price_entry.text())
            if not estimation_price:
                QMessageBox.warning(self.controller, "입력 오류", "메인 화면의 '추정가격'을 정확히 입력해주세요.")
                return

            base_amount = utils.parse_amount(self.controller.base_amount_entry.text())
            tuchal_amount_text = self.controller.tuchal_amount_label.text().replace("<b>", "").replace("</b>",
                                                                                                       "").replace(" 원",
                                                                                                                   "").replace(
                ",", "")
            tuchal_amount = utils.parse_amount(tuchal_amount_text) or 0

            price_data = {
                "estimation_price": estimation_price,
                "notice_base_amount": base_amount,
                "tuchal_amount": tuchal_amount
            }

            sipyung_info = {
                "is_limited": self.controller.sipyung_limit_check.isChecked(),
                "limit_amount": utils.parse_amount(self.controller.sipyung_limit_amount.text()) or 0,
                "method": "비율제" if self.controller.ratio_method_radio.isChecked() else "합산제",
                "tuchal_amount": price_data["tuchal_amount"]
            }

            calculation_context = {
                "announcement_date": announcement_date,
                "rule_info": rule_info,
                "price_data": price_data,
                "sipyung_info": sipyung_info,
                "region_limit": self.region_limit,
                "field_to_search": self.controller.gongo_field_combo.currentText()
            }
        except Exception as e:
            QMessageBox.critical(self, "오류", f"점수 계산에 필요한 공고 정보를 가져오는 데 실패했습니다:\n{e}")
            return
        # ▲▲▲▲▲ [핵심 수정] 여기까지 ▲▲▲▲▲

        initial_data_for_dialog = [widget.result_data for widget in self.result_widgets]

        dialog = ConsortiumManagerDialog(initial_data_for_dialog, self.region_limit, calculation_context, self)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            updated_consortiums_details_list = dialog.get_results()

            # ▼▼▼▼▼ [핵심 수정] 성적표(result_data)를 새로 발급(재계산)합니다 ▼▼▼▼▼
            if len(updated_consortiums_details_list) == len(self.result_widgets):

                context_for_calc = calculation_context.copy()
                context_for_calc.pop('field_to_search', None)  # 'field_to_search' 키를 안전하게 제거

                for i, new_details in enumerate(updated_consortiums_details_list):
                    if not new_details:  # 빈 협정은 건너뛰기
                        self.result_widgets[i].result_data['company_details'] = []
                        continue

                    # 변경된 업체 목록(new_details)과 공고 정보(context)로 점수를 재계산
                    recalculated_result = calculation_logic.calculate_consortium(
                        new_details, **context_for_calc
                    )

                    if recalculated_result:
                        # 기존 result_data에 공고 제목 등은 유지하면서 재계산된 결과를 덮어쓰기
                        current_gongo_title = self.result_widgets[i].result_data.get('gongo_title', '')
                        current_gongo_no = self.result_widgets[i].result_data.get('gongo_no', '')

                        self.result_widgets[i].result_data = recalculated_result
                        self.result_widgets[i].result_data['gongo_title'] = current_gongo_title
                        self.result_widgets[i].result_data['gongo_no'] = current_gongo_no
                    else:
                        # 계산 실패 시, 업체 목록만이라도 업데이트
                        self.result_widgets[i].result_data['company_details'] = new_details

            # ▲▲▲▲▲ [핵심 수정] 여기까지 ▲▲▲▲▲

            self.populate_consortium_list()
            self.update_detail_view()
            QMessageBox.information(self, "수정 완료", "전체 협정 정보가 성공적으로 수정되었습니다.")

            self.populate_consortium_list()
            self.update_detail_view()
            QMessageBox.information(self, "수정 완료", "전s체 협정 정보가 성공적으로 수정되었습니다.")