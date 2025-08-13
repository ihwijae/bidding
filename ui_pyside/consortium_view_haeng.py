# ui_pyside/consortium_view.py
import os
import sys
from PySide6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QGridLayout, QLabel, QLineEdit,
                               QPushButton, QTableWidget, QTableWidgetItem,
                               QHeaderView, QMessageBox, QFrame, QSplitter, QApplication, QScrollArea,
                               QComboBox, QDateEdit, QRadioButton, QGroupBox, QCheckBox, QMenu, QTextEdit)
from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QFont, QColor
import utils
import config
import calculation_logic # [핵심] 누락되었던 import 문
from .company_select_popup import CompanySelectPopupPyside
from .review_dialog import ReviewDialogPyside
from .guided_copy_popup import GuidedCopyPopup
from .share_check_popup import ShareCheckPopup
from .text_display_popup import TextDisplayPopup
# [파일 상단 import 구문에 아래 2개를 추가하세요]
from PySide6.QtCore import QDateTime # 날짜/시간 처리를 위해 추가
from .api_popup import ApiPopup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from PySide6.QtWidgets import QFileDialog
from .result_management_dialog import ResultManagementDialog
from .load_consortium_popup import LoadConsortiumPopup
import search_logic
import re


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


class  ConsortiumViewHaeng(QWidget):
    def __init__(self, controller):
        super().__init__(controller)
        self.controller = controller
        self.mode = "행안부"
        self.company_data_map = {}
        self.result_widgets = []
        self.announcement_date_modified = False
        self.bid_opening_date = None  # ▼▼▼▼▼ [추가] 개찰일자를 저장할 변수 ▼▼▼▼▼
        self.setup_ui()
        self.connect_signals()

    def setup_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)

        input_panel = self.create_input_panel()

        result_management_box = QGroupBox("계산 결과 요약")
        result_management_layout = QVBoxLayout(result_management_box)  # 레이아웃을 QVBoxLayout으로 변경

        # 상단 버튼 영역
        top_button_layout = QHBoxLayout()
        self.open_results_button = QPushButton("📂 결과 관리창 열기...")
        self.delete_all_button = QPushButton("🗑️ 전체 삭제")
        top_button_layout.addWidget(self.open_results_button)
        top_button_layout.addStretch(1)
        top_button_layout.addWidget(self.delete_all_button)

        # 요약 정보 표시 영역
        self.summary_display = QTextEdit()
        self.summary_display.setReadOnly(True)
        self.summary_display.setFont(QFont("맑은 고딕", 10))
        self.summary_display.setPlaceholderText("여기에 추가된 협정 결과가 요약되어 표시됩니다.")
        self.summary_display.setFixedHeight(100)  # 텍스트 상자 높이 고정

        self.summary_display.setLineWrapMode(QTextEdit.LineWrapMode.WidgetWidth)
        self.summary_display.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)

        result_management_layout.addLayout(top_button_layout)
        result_management_layout.addWidget(self.summary_display)  # 텍스트 상자 추가

        self.tree.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)


        main_layout.addWidget(input_panel)
        main_layout.addWidget(result_management_box)


    def connect_signals(self):
        self.api_search_button.clicked.connect(self.open_api_popup)
        self.estimation_price_entry.textChanged.connect(self.update_ui_by_rule)
        self.tree.doubleClicked.connect(self.on_tree_double_click)
        self.tree.itemChanged.connect(self.on_share_changed)
        self.review_button.clicked.connect(self.open_review_dialog)
        self.add_result_button.clicked.connect(self.process_and_add_result)
        self.delete_all_button.clicked.connect(self.delete_all_results)
        self.rule_combo.currentTextChanged.connect(self.update_ui_by_rule)
        self.announcement_date_edit.dateChanged.connect(self.on_announcement_date_changed)
        self.sipyung_limit_check.stateChanged.connect(self.on_sipyung_limit_toggled)
        self.estimation_price_entry.textChanged.connect(self.update_sipyung_limit_amount)
        self.base_amount_entry.textChanged.connect(self.calculate_tuchal_amount)
        self.tuchal_rate_entry.textChanged.connect(self.calculate_tuchal_amount)
        self.sajung_rate_entry.textChanged.connect(self.calculate_tuchal_amount)
        self.pre_check_button.clicked.connect(self.run_pre_check)
        self.tree.customContextMenuRequested.connect(self.show_context_menu)
        self.excel_export_button.clicked.connect(self.generate_excel_report)
        self.open_results_button.clicked.connect(self.open_result_management_dialog)
        self.load_button.clicked.connect(self.load_and_recalculate_consortium)


    def create_input_panel(self):
        panel = QWidget(); panel.setObjectName("filterBox"); layout = QVBoxLayout(panel)


        # [핵심 추가] 최상단에 버튼 추가
        api_button_layout = QHBoxLayout()
        self.api_search_button = QPushButton("📡 API 공고 검색")
        api_button_layout.addStretch(1)
        api_button_layout.addWidget(self.api_search_button)
        layout.addLayout(api_button_layout)
        
        notice_grid = QGridLayout()
        self.title_label = QLabel(); self.gongo_no_entry = QLineEdit(); self.gongo_title_entry = QLineEdit()
        self.estimation_price_entry = QLineEdit(); self.performance_target_label = QLabel()
        self.gongo_field_combo = QComboBox(); self.gongo_field_combo.addItem("-- 분야 선택 --"); self.gongo_field_combo.addItems(self.controller.source_files.keys())
        self.rule_combo = QComboBox()
        if self.mode in config.CONSORTIUM_RULES: self.rule_combo.addItems(config.CONSORTIUM_RULES[self.mode].keys())
        self.announcement_date_edit = QDateEdit(); self.announcement_date_edit.setCalendarPopup(True); self.announcement_date_edit.setDisplayFormat("yyyy-MM-dd"); self.announcement_date_edit.setDate(QDate.currentDate())
        self.performance_label = QLabel()
        notice_grid.addWidget(self.title_label, 0, 0, 1, 4)
        notice_grid.addWidget(QLabel("공고번호:"), 1, 0); notice_grid.addWidget(self.gongo_no_entry, 1, 1); notice_grid.addWidget(QLabel("<b>공고일:</b>"), 1, 2); notice_grid.addWidget(self.announcement_date_edit, 1, 3)
        notice_grid.addWidget(QLabel("공고제목:"), 2, 0); notice_grid.addWidget(self.gongo_title_entry, 2, 1); notice_grid.addWidget(QLabel("<b>공고분야:</b>"), 2, 2); notice_grid.addWidget(self.gongo_field_combo, 2, 3)
        notice_grid.addWidget(QLabel("<b>심사 기준:</b>"), 3, 0); notice_grid.addWidget(self.rule_combo, 3, 1, 1, 3)
        notice_grid.addWidget(QLabel("추정가격:"), 4, 0); notice_grid.addWidget(self.estimation_price_entry, 4, 1); notice_grid.addWidget(self.performance_label, 4, 2); notice_grid.addWidget(self.performance_target_label, 4, 3)
        layout.addLayout(notice_grid)

        self.bid_amount_group = QGroupBox("입찰금액 정보 (30억 이상)")
        bid_amount_layout = QGridLayout(self.bid_amount_group)
        
 # [핵심 수정] bid_amount_layout을 QGridLayout으로 변경하여 UI를 더 체계적으로 구성
        # --- 1행: 기초금액 ---
        self.base_amount_entry = QLineEdit()
        self.base_amount_entry.setPlaceholderText("기초금액을 입력하세요")
        bid_amount_layout.addWidget(QLabel("<b>기초금액:</b>"), 0, 0)
        bid_amount_layout.addWidget(self.base_amount_entry, 0, 1)

        # --- 2행: 투찰율 및 사정율 ---
        self.tuchal_rate_entry = QLineEdit()
        self.tuchal_rate_entry.setText("88.745") # 기본값 설정
        bid_amount_layout.addWidget(QLabel("<b>투찰율(%):</b>"), 1, 0)
        bid_amount_layout.addWidget(self.tuchal_rate_entry, 1, 1)

        self.sajung_rate_entry = QLineEdit()
        self.sajung_rate_entry.setText("101.8") # 기본값 설정
        bid_amount_layout.addWidget(QLabel("<b>사정율(%):</b>"), 1, 2)
        bid_amount_layout.addWidget(self.sajung_rate_entry, 1, 3)

        # --- 3행: 예상 투찰금액 ---
        self.tuchal_amount_label = QLabel("0 원")
        bid_amount_layout.addWidget(QLabel("<b>예상 투찰금액:</b>"), 2, 0)
        # 3행의 1열부터 3열까지 모두 차지하도록 setColumnStretch는 마지막에 설정
        bid_amount_layout.addWidget(self.tuchal_amount_label, 2, 1, 1, 3)
        
        # 열 너비 비율 조절
        bid_amount_layout.setColumnStretch(1, 1)
        bid_amount_layout.setColumnStretch(3, 1)

        layout.addWidget(self.bid_amount_group)

        qualification_group = QGroupBox("참가자격 제한"); qualification_layout = QGridLayout(qualification_group)
        self.region_limit_combo = QComboBox(); self.region_limit_combo.addItems(["전체", "서울", "경기", "인천", "강원", "충북", "충남", "대전", "세종", "전북", "전남", "광주", "경북", "경남", "대구", "울산", "부산", "제주"])
        self.duty_ratio_entry = QLineEdit(); self.duty_ratio_entry.setPlaceholderText("예: 49")
        qualification_layout.addWidget(QLabel("지역제한:"), 0, 0); qualification_layout.addWidget(self.region_limit_combo, 0, 1)
        qualification_layout.addWidget(QLabel("의무비율(%):"), 0, 2); qualification_layout.addWidget(self.duty_ratio_entry, 0, 3)
        self.sipyung_limit_check = QCheckBox("시평액 제한 있음")
        self.sipyung_limit_amount = QLineEdit(); self.sipyung_limit_amount.setPlaceholderText("제한 금액(추정가격 기준)"); self.sipyung_limit_amount.setEnabled(False)
        self.ratio_method_radio = QRadioButton("비율제"); self.ratio_method_radio.setChecked(True)
        self.sum_method_radio = QRadioButton("합산제")
        self.ratio_method_radio.setEnabled(False); self.sum_method_radio.setEnabled(False)
        method_layout = QHBoxLayout(); method_layout.addWidget(self.ratio_method_radio); method_layout.addWidget(self.sum_method_radio); method_layout.addStretch()
        qualification_layout.addWidget(self.sipyung_limit_check, 1, 0); qualification_layout.addWidget(self.sipyung_limit_amount, 1, 1)
        qualification_layout.addWidget(QLabel("계산방식:"), 1, 2); qualification_layout.addLayout(method_layout, 1, 3)
        layout.addWidget(qualification_group)

 # [이 코드를 추가하세요]
        # --- 업체 구성 테이블 ---
        self.tree = QTableWidget()
        self.tree.setRowCount(5)
        self.tree.setColumnCount(5)
        self.tree.setHorizontalHeaderLabels(["구분", "업체명", "지역", "5년실적", "지분율(%)"])
        self.tree.verticalHeader().setVisible(False)
        roles = ["대표사"] + [f"구성사 {i}" for i in range(1, 5)]
        for i, role in enumerate(roles):
            item_role = QTableWidgetItem(role)
            item_role.setFlags(item_role.flags() & ~Qt.ItemIsEditable)
            self.tree.setItem(i, 0, item_role)
            item_name = QTableWidgetItem("[더블클릭하여 업체 선택]")
            item_name.setFlags(item_name.flags() & ~Qt.ItemIsEditable)
            self.tree.setItem(i, 1, item_name)
            for c in range(2, 5):
                item = QTableWidgetItem("")
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.tree.setItem(i, c, item)
            self.company_data_map[i] = {'role': role, 'data': None, 'share': 0, 'source_type': None}
        self.tree.resizeColumnsToContents()
        self.tree.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.tree, 1)
        
        # --- [핵심 수정] 하단 버튼 레이아웃 ---
        button_layout = QHBoxLayout()

        self.load_button = QPushButton("📂 협정 불러오기")
        button_layout.addWidget(self.load_button)

        self.excel_export_button = QPushButton("💾 엑셀로 저장") # 새 버튼 생성
        button_layout.addWidget(self.excel_export_button)
        button_layout.addStretch(1) # 버튼들을 오른쪽으로 밀어줌

        # 1. 지분율 사전검토 버튼
        self.pre_check_button = QPushButton("🔬 지분율 사전검토")
        button_layout.addWidget(self.pre_check_button)
        
        # 2. 적격심사 검토 버튼
        self.review_button = QPushButton("📋 적격심사 검토")
        button_layout.addWidget(self.review_button)
        
        # 3. 결과 표 추가 버튼
        self.add_result_button = QPushButton("📊 결과 표 추가")
        button_layout.addWidget(self.add_result_button)
        
        layout.addLayout(button_layout)
        
        self.update_ui_by_rule() # 마지막에 호출
        return panel      


    def create_result_scroll_area(self):
        container_widget = QWidget(); container_layout = QVBoxLayout(container_widget)
        container_layout.setContentsMargins(0, 0, 0, 0)
        top_button_layout = QHBoxLayout(); top_button_layout.addWidget(QLabel("<b>계산 결과 목록</b>"))
        top_button_layout.addStretch(1)

        self.generate_messages_button = QPushButton("✉️ 협정 문자 일괄 생성")
        top_button_layout.addWidget(self.generate_messages_button)

        self.delete_all_button = QPushButton("🗑️ 전체 삭제")
        top_button_layout.addWidget(self.delete_all_button); container_layout.addLayout(top_button_layout)
        scroll_area = QScrollArea(); scroll_area.setWidgetResizable(True); scroll_area.setObjectName("filterBox")
        scroll_content = QWidget(); self.results_layout = QVBoxLayout(scroll_content)
        self.results_layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        scroll_area.setWidget(scroll_content); container_layout.addWidget(scroll_area)
        return container_widget

    def update_ui_by_rule(self):
        selected_rule_key = self.rule_combo.currentText()
        if not selected_rule_key: return
        try:
            ruleset = config.CONSORTIUM_RULES[self.mode][selected_rule_key]
            self.title_label.setText(f"<b>공고 정보 ({ruleset.get('name', self.mode)})</b>")
            multiplier = ruleset.get('performance_multiplier', 1.0)
            self.performance_label.setText(f"실적만점({multiplier}배수):")
            if "30억이상" in selected_rule_key: self.bid_amount_group.setVisible(True)
            else: self.bid_amount_group.setVisible(False); self.base_amount_entry.clear()
            self.calculate_performance_target()
        except KeyError:
            self.bid_amount_group.setVisible(False); self.base_amount_entry.clear()
    
    def calculate_performance_target(self):
        text = self.estimation_price_entry.text()
        selected_rule_key = self.rule_combo.currentText()
        ruleset = config.CONSORTIUM_RULES.get(self.mode, {}).get(selected_rule_key, {})
        multiplier = ruleset.get('performance_multiplier', 1.0)
        price_val = utils.parse_amount(text)
        target = price_val * multiplier if price_val else 0
        self.performance_target_label.setText(f"{target:,.0f}" if target else "")
        
    
    def calculate_tuchal_amount(self):
        try:
            # 1. 모든 입력칸에서 텍스트를 읽어와 숫자로 변환 (실패 시 0으로 처리)
            base_amount = utils.parse_amount(self.base_amount_entry.text()) or 0
            tuchal_rate = float(self.tuchal_rate_entry.text()) or 0
            sajung_rate = float(self.sajung_rate_entry.text()) or 0
        except (ValueError, TypeError):
            # 입력칸에 숫자가 아닌 문자가 들어올 경우를 대비
            self.tuchal_amount_label.setText("<b style='color:red;'>숫자만 입력하세요</b>")
            return

        # 2. 모든 값이 유효할 경우에만 계산
        if base_amount > 0 and tuchal_rate > 0 and sajung_rate > 0:
            tuchal_amount = base_amount * (tuchal_rate / 100.0) * (sajung_rate / 100.0)
            self.tuchal_amount_label.setText(f"<b style='color:blue;'>{tuchal_amount:,.0f} 원</b>")
        else:
            self.tuchal_amount_label.setText("0 원")

    def on_announcement_date_changed(self):
        self.announcement_date_modified = True

    def on_sipyung_limit_toggled(self, state):
        is_enabled = (state == Qt.CheckState.Checked.value)
        self.sipyung_limit_amount.setEnabled(is_enabled)
        self.ratio_method_radio.setEnabled(is_enabled)
        self.sum_method_radio.setEnabled(is_enabled)
        if is_enabled: self.update_sipyung_limit_amount()
        else: self.sipyung_limit_amount.clear()

    def update_sipyung_limit_amount(self):
        if self.sipyung_limit_check.isChecked():
            estimation_price_text = self.estimation_price_entry.text()
            self.sipyung_limit_amount.setText(estimation_price_text)

    def validate_inputs(self):
        # --- 상단부 유효성 검사는 기존과 동일 ---
        if not self.announcement_date_modified:
            QMessageBox.warning(self, "입력 필요", "정확한 계산을 위해 '공고일'을 반드시 설정해주세요.");
            return None
        announcement_date = self.announcement_date_edit.date().toPython()

        companies_data = [info for i, info in self.company_data_map.items() if
                          info and info['data'] and info.get('share', 0) > 0]
        if not companies_data:
            QMessageBox.warning(self, "입력 오류", "업체를 1곳 이상 선택하고, 지분율을 0보다 크게 입력하세요.");
            return None
        selected_rule_key = self.rule_combo.currentText()
        if not selected_rule_key:
            QMessageBox.warning(self, "선택 오류", "심사 기준을 선택하세요.");
            return None
        rule_info = (self.mode, selected_rule_key)
        region_limit = self.region_limit_combo.currentText()

        # ... (지역제한, 의무비율 검사 로직은 기존과 동일) ...

        # ▼▼▼▼▼ [핵심 수정] 모든 금액 정보를 price_data 딕셔너리로 묶기 ▼▼▼▼▼

        # 1. 추정가격 검증
        estimation_price_val = utils.parse_amount(self.estimation_price_entry.text())
        if not estimation_price_val:
            QMessageBox.warning(self, "입력 오류", "추정가격을 정확히 입력해주세요.");
            return None

        # 2. 기초금액 가져오기 (행안부/조달청 UI 자동 호환)
        base_amount_val = 0
        if hasattr(self, 'notice_base_amount_entry'):  # 조달청 UI에 해당 위젯이 있으면
            base_amount_val = utils.parse_amount(self.notice_base_amount_entry.text())
        elif hasattr(self, 'base_amount_entry'):  # 행안부 UI에 해당 위젯이 있으면
            base_amount_val = utils.parse_amount(self.base_amount_entry.text())

        # 3. 투찰금액 가져오기
        tuchal_amount_text = self.tuchal_amount_label.text().replace("<b>", "").replace("</b>", "").replace(" 원",
                                                                                                            "").replace(
            ",", "")
        tuchal_amount_val = utils.parse_amount(tuchal_amount_text) or 0

        # 4. 모든 금액 정보를 하나의 딕셔너리로 생성
        price_data = {
            "estimation_price": estimation_price_val,
            "notice_base_amount": base_amount_val,
            "tuchal_amount": tuchal_amount_val
        }

        # ▲▲▲▲▲ [핵심 수정] 여기까지 ▲▲▲▲▲

        sipyung_info = {
            "is_limited": self.sipyung_limit_check.isChecked(),
            "limit_amount": utils.parse_amount(self.sipyung_limit_amount.text()) or 0,
            "method": "비율제" if self.ratio_method_radio.isChecked() else "합산제",
            "tuchal_amount": price_data["tuchal_amount"]  # 위에서 계산한 값을 사용
        }

        # [수정] 반환값에서 estimation_price 대신 price_data를 반환
        return (companies_data, price_data, announcement_date, rule_info, sipyung_info, region_limit)

    def open_review_dialog(self):
        validated_data = self.validate_inputs()
        if not validated_data: return

        companies_data, price_data, announcement_date, rule_info, sipyung_info, region_limit = validated_data

        if not self.check_regional_requirements(companies_data):
            return

        result = calculation_logic.calculate_consortium(companies_data, price_data, announcement_date, rule_info,
                                                        sipyung_info, region_limit)

        if not result:
            QMessageBox.critical(self, "계산 오류", "점수 계산 중 오류가 발생했습니다.")
            return

        review_dialog = ReviewDialogPyside(result, self)
        review_dialog.exec()

    def process_and_add_result(self):
        validated_data = self.validate_inputs()
        if not validated_data: return

        # validate_inputs가 반환한 price_data를 올바르게 받습니다.
        companies_data, price_data, announcement_date, rule_info, sipyung_info, region_limit = validated_data

        # 지역 요건 검사를 먼저 수행합니다.
        if not self.check_regional_requirements(companies_data):
            return

        current_company_names = {comp['data'].get("검색된 회사") for comp in companies_data}

        # 중복 검사 로직 (기존과 동일)
        existing_company_names = set()
        for result_widget in self.result_widgets:
            if hasattr(result_widget, 'result_data'):
                for detail in result_widget.result_data.get("company_details", []):
                    existing_company_names.add(detail.get("name"))

        overlapping_companies = current_company_names.intersection(existing_company_names)

        if overlapping_companies:
            names_str = ", ".join(overlapping_companies)
            QMessageBox.critical(self, "중복 오류",
                                 f"이미 다른 결과에 포함된 업체가 있습니다: [{names_str}]\n\n"
                                 "기존 결과를 삭제하거나 다른 업체로 구성해주세요.")
            return

        # calculation_logic 호출 시 price_data를 전달합니다.
        result = calculation_logic.calculate_consortium(companies_data, price_data, announcement_date, rule_info,
                                                        sipyung_info, region_limit)


        if not result:
            QMessageBox.critical(self, "계산 오류", "점수 계산 중 오류가 발생했습니다.")
            return

        result['gongo_title'] = self.gongo_title_entry.text()
        result['gongo_no'] = self.gongo_no_entry.text()

        result_index = len(self.result_widgets)
        result_widget = self.create_single_result_widget(result, result_index)

        # self.results_layout.addWidget(result_widget)
        self.result_widgets.append(result_widget)
        self.update_summary_display()

    # [on_tree_double_click 함수를 이 코드로 통째로 교체하세요]
    def on_tree_double_click(self, model_index):
        row = model_index.row()
        col = model_index.column()
        if col != 1: return

        selected_field = self.gongo_field_combo.currentText()
        if selected_field == "-- 분야 선택 --":
            QMessageBox.warning(self, "선택 필요", "먼저 '공고분야'를 선택해주세요."); return

        # [핵심] 현재 테이블에 추가된 업체들의 이름 목록을 다시 정확하게 생성합니다.
        existing_companies = []
        for r in range(self.tree.rowCount()):
            # 더블클릭한 자기 자신 행은 중복 검사에서 제외해야 합니다.
            if r == row: continue
            
            # company_data_map에서 업체 정보를 가져옵니다.
            company_info = self.company_data_map.get(r)
            if company_info and company_info.get('data'):
                company_name = company_info['data'].get("검색된 회사")
                if company_name:
                    existing_companies.append(company_name)
        
        # --- 디버깅용 print문 (문제가 해결되면 지워도 됩니다) ---
        print(f"중복 검사를 위해 팝업으로 전달하는 업체 목록: {existing_companies}")
        
        # --- 콜백 함수 정의 ---
        def update_company_info(selected_company_data):
            self.tree.blockSignals(True)
            self.company_data_map[row]['data'] = selected_company_data
            self.company_data_map[row]['source_type'] = selected_field
            company_name = selected_company_data.get("검색된 회사", "")
            region = selected_company_data.get("지역", "")
            perf_5y = utils.parse_amount(selected_company_data.get("5년 실적"))
            self.tree.setItem(row, 1, QTableWidgetItem(company_name))
            self.tree.setItem(row, 2, QTableWidgetItem(region))
            self.tree.setItem(row, 3, QTableWidgetItem(f"{perf_5y:,.0f}" if perf_5y is not None else "0"))
            self.tree.blockSignals(False)
            self.tree.setItem(row, 4, QTableWidgetItem("0"))
            self.tree.resizeColumnsToContents()
            self.tree.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
            self.pre_check_passed = False

        # --- 팝업창 생성 및 실행 ---
        self.popup = CompanySelectPopupPyside(self, self.controller, selected_field, update_company_info, existing_companies)
        self.popup.show()

    def on_share_changed(self, item):
        if item.column() == 4:  # 지분율(%) 열
            try:
                # 사용자가 입력한 텍스트를 100으로 나누어 0.51과 같은 형태로 저장
                share_value = float(item.text()) / 100.0
                self.company_data_map[item.row()]['share'] = share_value
            except (ValueError, TypeError):
                self.company_data_map[item.row()]['share'] = 0
            
    def delete_all_results(self):
        reply = QMessageBox.question(self, "전체 삭제", "모든 계산 결과를 삭제하시겠습니까?")
        if reply == QMessageBox.StandardButton.Yes:
            for widget in self.result_widgets: widget.deleteLater()
            self.result_widgets.clear()

            self.update_summary_display()
            
    # [delete_single_result 함수를 이 코드로 통째로 교체하세요]
    def delete_single_result(self, widget_to_delete):
        # 1. 사용자에게 삭제 여부를 먼저 확인합니다.
        reply = QMessageBox.question(self, "결과 삭제", "선택한 계산 결과를 삭제하시겠습니까?",
                                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                                    QMessageBox.StandardButton.No)

        if reply != QMessageBox.StandardButton.Yes:
            return

        # 2. 'Yes'를 눌렀을 때만, 내부 삭제 함수를 호출합니다.
        self._remove_widget_from_list(widget_to_delete)
        
    # [create_single_result_widget 함수를 이 코드로 통째로 교체하세요]
    def create_single_result_widget(self, result_data, index):
        frame = QFrame()
        frame.result_data = result_data
        frame.setFrameShape(QFrame.Shape.StyledPanel)
        frame.setLineWidth(1)
        layout = QVBoxLayout(frame)
        top_layout = QHBoxLayout()
        notice_info_label = QLabel(f"<b>#{index + 1} | 공고:</b> {result_data.get('gongo_title', '없음')}")
        top_layout.addWidget(notice_info_label)

        # ▼▼▼▼▼ [수정] top_layout.add Stretch(1)을 아래와 같이 한 줄로 수정 ▼▼▼▼▼
        top_layout.addStretch(1)

        # [핵심] 수정 버튼 추가
        edit_button = QPushButton("✏️ 수정")
        edit_button.clicked.connect(lambda _, r=result_data, w=frame: self.edit_result(r, w))
        top_layout.addWidget(edit_button)

        copy_button = QPushButton("엑셀에 안전 복사")
        report_table = QTableWidget()
        copy_button.clicked.connect(lambda: self.start_guided_copy(report_table, index))

        delete_button = QPushButton("X")
        delete_button.setFixedWidth(30)
        delete_button.clicked.connect(lambda: self.delete_single_result(frame))

        top_layout.addWidget(copy_button)
        top_layout.addWidget(delete_button)
        layout.addLayout(top_layout)

        self.populate_report_table(report_table, result_data)
        layout.addWidget(report_table)

        status_label = QLabel()
        status_label.setStyleSheet("color: blue; font-size: 11px; padding-left: 5px;")
        credit_applied = any(
            d['business_score_details'].get('basis') == '신용평가' for d in result_data.get('company_details', []))
        if credit_applied:
            status_label.setText("<b>(*) 일부 경영점수는 신용평가 기준으로 산정되었습니다.</b>")
        layout.addWidget(status_label)

        return frame

    # [populate_report_table 함수를 이 코드로 통째로 교체하세요]
    def populate_report_table(self, table, result_data):
        # --- 1. 테이블 기본 설정 ---
        table.clear()
        table.setColumnCount(32) 
        table.setRowCount(3)
        table.verticalHeader().setVisible(False)
        table.horizontalHeader().setVisible(False)
        
        # [핵심] '업체명' 열(0~5)의 너비를 150으로, 나머지는 60으로 설정
        for i in range(6):
            table.setColumnWidth(i, 150) # 기존 100 -> 150으로 대폭 증가
        for i in range(6, 32):
            table.setColumnWidth(i, 60) # 나머지는 60 유지
            
        for i in range(3):
            table.setRowHeight(i, 28)
        table.setFixedHeight(table.horizontalHeader().height() + table.rowHeight(0) * 3)

        # --- 2. 헤더 생성 (기존과 동일) ---
        self.merge_and_set_item(table, 0, 0, 1, 6, "업체명")
        self.merge_and_set_item(table, 0, 6, 1, 7, "지분")
        self.merge_and_set_item(table, 0, 13, 1, 6, "경영상태")
        self.merge_and_set_item(table, 0, 19, 1, 7, "시공실적")
        self.merge_and_set_item(table, 0, 26, 1, 6, "시공능력") 

        headers2_base = ["대표사", "구성원1", "구성원2", "구성원3", "구성원4", "비고", "대표사", "구성원1", "구성원2", "구성원3", "구성원4", "합산지분", "가점", "대표사", "구성원1", "구성원2", "구성원3", "구성원4", "경영(15)", "대표사", "구성원1", "구성원2", "구성원3", "구성원4", "실적비율", "실적점수"]
        headers2_sipyung = ["시공능력 대표사", "시공능력 구성사1", "시공능력 구성사2", "시공능력 구성사3", "시공능력 구성사4", "시공능력+시공비율"]
        headers2 = headers2_base + headers2_sipyung
        for c, h in enumerate(headers2):
            self.merge_and_set_item(table, 1, c, 1, 1, h)

        # --- 3. 데이터 채우기 (기존과 동일) ---
        details = result_data.get("company_details", [])
        data_row = 2
        for comp_detail in details:
            role = comp_detail.get('role'); col_offset = 0
            if role == "대표사": col_offset = 0
            elif role.startswith("구성사"):
                try: col_offset = int(role.split(' ')[1])
                except: continue
            business_score = comp_detail.get('business_score_details', {}).get('total', 0)
            performance_5y = comp_detail.get('performance_5y', 0)

            # ▼▼▼▼▼ [디버깅] 지분율 값을 단계별로 출력 ▼▼▼▼▼
            print(f"\n--- [디버깅] populate_report_table 함수 ---")
            share_decimal = comp_detail.get('share', 0)
            print(f"[1] result_data에서 가져온 share 값 (소수): {share_decimal}")

            share_percent = share_decimal * 100.0
            print(f"[2] 100을 곱한 퍼센트 값: {share_percent}")

            formatted_share_text = f"{share_percent:.1f}%"
            print(f"[3] 최종적으로 표에 표시될 텍스트: '{formatted_share_text}'")
            print("------------------------------------")
            # ▲▲▲▲▲ [디버깅] 여기까지 ▲▲▲▲▲


            self.set_item(table, data_row, col_offset, comp_detail.get('name', ''))

            # ▼▼▼▼▼ [핵심 수정] 지분율을 100 곱해서 퍼센트로 표시 ▼▼▼▼▼
            share_decimal = comp_detail.get('share', 0)  # 1.0, 0.51 같은 소수점 값
            share_percent = share_decimal * 100.0  # 100, 51과 같은 퍼센트 값
            self.set_item(table, data_row, col_offset + 6, f"{share_percent:.1f}%")
            # ▲▲▲▲▲ [핵심 수정] 여기까지 ▲▲▲▲▲

            self.set_item(table, data_row, col_offset + 13, f"{business_score:.4f}")
            self.set_item(table, data_row, col_offset + 19, f"{performance_5y:,.0f}" if performance_5y else "0")
            sipyung_amount = utils.parse_amount(str(comp_detail['data'].get("시평", 0))) or 0
            self.set_item(table, data_row, col_offset + 26, f"{sipyung_amount:,.0f}" if sipyung_amount else "0")

        final_biz_score = result_data.get('final_business_score', 0)
        biz_score_item = self.set_item(table, data_row, 18, f"{final_biz_score:.4f}")
        if abs(final_biz_score - 15.0) > 0.001: biz_score_item.setForeground(QColor("red"))

        # --- 4. 스타일 적용 (기존과 동일) ---
        header_font = QFont(); header_font.setBold(True); header_bg_color = QColor("#F0F0F0")
        for r in range(2):
            for c in range(32):
                item = table.item(r, c)
                if item: item.setFont(header_font); item.setBackground(header_bg_color); item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        for c in range(32):
            item = table.item(data_row, c)
            if item: item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
    
    def set_item(self, table, r, c, text):
        item = QTableWidgetItem(str(text)); table.setItem(r, c, item); return item
    def merge_and_set_item(self, table, r, c, r_span, c_span, text):
        if r_span > 1 or c_span > 1: table.setSpan(r, c, r_span, c_span)
        self.set_item(table, r, c, text)
        
    # [start_guided_copy 함수를 이 코드로 통째로 교체하세요]
    def start_guided_copy(self, table, result_index):
        excel_start_row = 3 + result_index
        
        # [핵심 수정] 복사할 데이터 덩어리(chunks)에 '시공능력' 부분 추가
        chunks = [
            {"name": "업체명/지분", "cols": range(0, 11), "start_cell": f"A{excel_start_row}"}, 
            {"name": "경영상태", "cols": range(13, 18), "start_cell": f"N{excel_start_row}"}, 
            {"name": "시공실적", "cols": range(19, 24), "start_cell": f"T{excel_start_row}"},
            # 새로운 '시공능력' 덩어리 추가. 26번 열부터 31번 열까지.
            # 마지막 "시공능력+시공비율" 열은 비워둬야 하므로, 26~30번 열(5개)만 복사.
            {"name": "시공능력", "cols": range(26, 31), "start_cell": f"AA{excel_start_row}"} 
        ]
        
        copy_chunks_data = []
        for chunk in chunks:
            row_items = []
            for c in chunk["cols"]:
                item = table.item(2, c)
                row_items.append(item.text() if item else "")
            
            # 팝업창에 표시될 안내 메시지 개선
            instruction = f"<b>{chunk['name']}</b> 데이터를 엑셀의 <b>{chunk['start_cell']}</b> 셀에 붙여넣으세요."
            copy_chunks_data.append({"instruction": instruction, "data": "\t".join(row_items)})
            
        self.guided_popup = GuidedCopyPopup(copy_chunks_data, self)
        self.guided_popup.exec()

    def run_pre_check(self):
        """'지분율 사전검토' 버튼 클릭 시 실행됩니다."""

        # 1. '30억 이상' 규칙일 때만 동작하도록 검증
        selected_rule_key = self.rule_combo.currentText()
        if "30억이상" not in selected_rule_key:
            QMessageBox.information(self, "알림", "이 기능은 '30억 이상' 심사 기준에만 적용됩니다.")
            return

        # 2. 투찰금액이 있는지 확인
        tuchal_amount_text = self.tuchal_amount_label.text().replace("<b>", "").replace("</b>", "").replace(" 원",
                                                                                                            "").replace(
            ",", "")
        tuchal_amount = utils.parse_amount(tuchal_amount_text) or 0
        if tuchal_amount <= 0:
            QMessageBox.warning(self, "입력 필요", "'기초금액'을 입력하여 투찰금액을 먼저 계산해주세요.")
            return

        # 3. 참여 업체 정보 수집
        companies_data = [info for i, info in self.company_data_map.items() if
                          info and info['data'] and info.get('share', 0) > 0]
        if not companies_data:
            QMessageBox.warning(self, "입력 필요", "검토할 업체를 먼저 선택하고 지분율을 입력해주세요.")
            return

        # ▼▼▼▼▼ [핵심 추가] 지역 의무비율 검사 ▼▼▼▼▼
        if not self.check_regional_requirements(companies_data):
            return  # 사용자가 '아니오'를 누르면 여기서 중단
        # ▲▲▲▲▲ 여기까지 추가 ▲▲▲▲▲

        # 4. 실제 계산 로직 호출
        results = calculation_logic.check_share_limit(companies_data, tuchal_amount)

        # 5. 통과 여부 상태 업데이트
        if any(res['is_problem'] for res in results):
            self.pre_check_passed = False
        else:
            self.pre_check_passed = True
            QMessageBox.information(self, "사전검토 통과", "모든 업체의 지분율이 참여 가능 한도 내에 있습니다.")

        # 6. 결과 팝업창 표시
        popup = ShareCheckPopup(results, self)
        popup.exec()

    # [이 함수 두 개를 클래스 내부에 추가하세요]

    def show_context_menu(self, pos):
        """테이블에서 우클릭 시 컨텍스트 메뉴를 표시합니다."""
        item = self.tree.itemAt(pos)
        if not item:
            return  # 빈 공간을 클릭하면 메뉴를 띄우지 않음

        row_index = item.row()
        # 이미 비어있는 행에는 '제거' 메뉴를 보여줄 필요 없음
        if not self.company_data_map[row_index]['data']:
            return

        menu = QMenu()
        remove_action = menu.addAction("선택한 업체 제거")
        
        # '제거' 액션을 클릭하면 remove_selected_company 함수를 실행
        action = menu.exec(self.tree.viewport().mapToGlobal(pos))
        
        if action == remove_action:
            self.remove_selected_company(row_index)

    # [remove_selected_company 함수를 이 코드로 통째로 교체하세요]
    def remove_selected_company(self, row_index, silent=False):
        """
        선택된 행의 업체 정보를 초기화합니다.
        silent=True이면, 완료 메시지를 띄우지 않습니다.
        """
        # 1. UI 테이블의 내용을 초기 상태로 되돌립니다.
        self.tree.blockSignals(True) # itemChanged 시그널이 불필요하게 실행되는 것을 방지
        
        self.tree.setItem(row_index, 1, QTableWidgetItem("[더블클릭하여 업체 선택]"))
        self.tree.setItem(row_index, 2, QTableWidgetItem("")) # 지역
        self.tree.setItem(row_index, 3, QTableWidgetItem("")) # 5년실적
        self.tree.setItem(row_index, 4, QTableWidgetItem("0"))  # 지분율
        
        self.tree.blockSignals(False)

        # 2. 내부 데이터 저장소(map)의 정보도 초기화합니다.
        self.company_data_map[row_index]['data'] = None
        self.company_data_map[row_index]['share'] = 0
        self.company_data_map[row_index]['source_type'] = None

        # 3. 컨소시엄 구성이 변경되었으므로, 사전검토 상태를 리셋합니다.
        self.pre_check_passed = False

        # 4. silent 모드가 아닐 때만 완료 메시지를 표시합니다.
        if not silent:
            QMessageBox.information(self, "알림", f"{row_index + 1}번째 행의 업체 정보가 초기화되었습니다.")



    def open_api_popup(self):
        # ApiPopup을 생성하고, gongo_selected 시그널을 fill_gongo_data 슬롯에 연결
        popup = ApiPopup(self)
        popup.gongo_selected.connect(self.fill_gongo_data)
        popup.exec()

        # [fill_gongo_data 함수를 이 코드로 통째로 교체하세요]
    def fill_gongo_data(self, gongo_data):
        # 1. API 데이터에서 필요한 모든 값을 안전하게 추출합니다.
        self.gongo_no_entry.setText(f"{gongo_data.get('bidNtceNo', '')}-{gongo_data.get('bidNtceOrd', '')}")
        self.gongo_title_entry.setText(gongo_data.get('bidNtceNm', ''))
        
        estimation_price_str = gongo_data.get('mainCnsttyPresmptPrce', gongo_data.get('presmptPrce', '0'))
        base_price_str = gongo_data.get('mainCnsttyCnstwkPrearngAmt', gongo_data.get('bssamt', '0'))
        cnstty_name = gongo_data.get('mainCnsttyNm', '')
        region_name_full = gongo_data.get('jntcontrctDutyRgnNm1', '')
        duty_rate_str = gongo_data.get('rgnDutyJntcontrctRt', '')
        
        # [핵심] 공고일 키를 'ntceDt'에서 'rgstDt'로 변경
        # "2025-07-04 09:09:26" 형식에서 날짜 부분만 사용
        ntce_dt_str = gongo_data.get('rgstDt', '')

        # ▼▼▼▼▼ [추가] API 응답에서 개찰일자(opengDt) 가져오기 ▼▼▼▼▼
        opening_dt_str = gongo_data.get('opengDt', '')
        if opening_dt_str:
            self.bid_opening_date = QDateTime.fromString(opening_dt_str.split('.')[0], "yyyy-MM-dd HH:mm:ss")
        # ▲▲▲▲▲ [추가] 여기까지 ▲▲▲▲▲

        # 2. 추출한 데이터로 UI 위젯의 값을 채웁니다.
        # 추정가격 설정
        try:
            price_val = int(float(estimation_price_str))
            self.estimation_price_entry.setText(f"{price_val:,}")
        except (ValueError, TypeError):
            self.estimation_price_entry.setText(estimation_price_str)
            
        # 공고일 설정
        if ntce_dt_str:
            # "2025-07-04 09:09:26" -> "2025-07-04"
            date_part = ntce_dt_str.split(' ')[0]
            q_date = QDate.fromString(date_part, "yyyy-MM-dd")
            if q_date.isValid():
                self.announcement_date_edit.setDate(q_date)

        # 분야 자동 선택
        if "전기" in cnstty_name: self.gongo_field_combo.setCurrentText("전기")
        elif "정보통신" in cnstty_name: self.gongo_field_combo.setCurrentText("통신")
        elif "소방" in cnstty_name: self.gongo_field_combo.setCurrentText("소방")
        else: self.gongo_field_combo.setCurrentText("기타")

        # 지역제한 자동 선택
        region_map = {"서울특별시": "서울", "경기도": "경기", "인천광역시": "인천", "강원특별자치도": "강원", "충청북도": "충북", "충청남도": "충남", "대전광역시": "대전", "세종특별자치시": "세종", "전북특별자치도": "전북", "전라남도": "전남", "광주광역시": "광주", "경상북도": "경북", "경상남도": "경남", "대구광역시": "대구", "울산광역시": "울산", "부산광역시": "부산", "제주특별자치도": "제주"}
        short_region_name = region_map.get(region_name_full, "전국")
        self.region_limit_combo.setCurrentText(short_region_name)
        
        # 의무비율 자동 입력
        if duty_rate_str:
            self.duty_ratio_entry.setText(duty_rate_str)
            
        # 기초금액 자동 입력
        if base_price_str:
            try:
                base_val = int(float(base_price_str))
                self.base_amount_entry.setText(f"{base_val:,}")
            except (ValueError, TypeError):
                self.base_amount_entry.setText(base_price_str)
                
        QMessageBox.information(self, "정보 입력 완료", "API 공고 정보가 자동으로 입력되었습니다.")


        # [ConsortiumViewHaeng 클래스 내부에 이 함수를 통째로 추가하세요]
    def edit_result(self, result_data, widget_to_edit):
        reply = QMessageBox.question(self, "결과 수정", 
                                    "이 결과를 상단 입력창으로 불러와 수정하시겠습니까?\n"
                                    "수정 후 '결과 표 추가' 버튼을 다시 누르면 이 표가 업데이트됩니다.",
                                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        
        if reply == QMessageBox.StandardButton.No:
            return

        # 1. 공고 정보 불러오기 (result_data에 저장된 값을 사용)
        self.gongo_title_entry.setText(result_data.get('gongo_title', ''))
        self.gongo_no_entry.setText(result_data.get('gongo_no', ''))
        # 다른 공고 정보(추정가격, 공고일 등)는 result_data에 없으므로,
        # 사용자가 다시 설정해야 합니다. 이 부분은 현재 로직을 유지합니다.
        
        # 2. 업체 구성 불러오기
        # 먼저 현재 입력 테이블 초기화
        for i in range(self.tree.rowCount()):
            self.remove_selected_company(i, silent=True)
            
        details = result_data.get("company_details", [])
        for comp_detail in details:
            role = comp_detail.get('role')
            # 역할(대표사/구성사)에 따라 테이블의 행 인덱스를 결정
            row_index = 0 
            if role.startswith("구성사"):
                try:
                    row_index = int(role.split(' ')[1])
                except (ValueError, IndexError):
                    continue
            
            # 테이블의 해당 행에 업체 정보를 다시 채워넣음
            self.tree.blockSignals(True)
            self.company_data_map[row_index]['data'] = comp_detail['data']
            self.company_data_map[row_index]['share'] = comp_detail['share']
            company_name = comp_detail.get("name", "")
            region = comp_detail['data'].get("지역", "")
            perf_5y = utils.parse_amount(comp_detail['data'].get("5년 실적"))
            self.tree.setItem(row_index, 1, QTableWidgetItem(company_name))
            self.tree.setItem(row_index, 2, QTableWidgetItem(region))
            self.tree.setItem(row_index, 3, QTableWidgetItem(f"{perf_5y:,.0f}" if perf_5y is not None else "0"))
            self.tree.setItem(row_index, 4, QTableWidgetItem(str(comp_detail['share'])))
            self.tree.blockSignals(False)

        # 3. 기존 결과 위젯 삭제
        self._remove_widget_from_list(widget_to_edit)
        
        QMessageBox.information(self, "불러오기 완료", "선택한 결과가 상단 입력창에 로드되었습니다.\n수정 후 '결과 표 추가' 버튼을 눌러주세요.")

        # [ConsortiumViewHaeng 클래스 내부에 이 함수를 통째로 추가하세요]
    def _remove_widget_from_list(self, widget_to_delete):
        """확인창 없이 위젯을 화면과 리스트에서 제거합니다."""
        widget_to_delete.deleteLater()
        if widget_to_delete in self.result_widgets:
            self.result_widgets.remove(widget_to_delete)

    def check_regional_requirements(self, companies_data):
        """지역제한 및 의무비율을 검사하고, 미충족 시 사용자에게 계속 진행할지 묻습니다."""
        region_limit = self.region_limit_combo.currentText()
        duty_ratio_str = self.duty_ratio_entry.text().strip()

        if region_limit != "전체" and duty_ratio_str:
            try:
                duty_ratio = float(duty_ratio_str)  # 사용자가 입력한 의무비율 (예: 49.0)
            except ValueError:
                QMessageBox.warning(self, "입력 오류", "'의무비율'에는 숫자만 입력해주세요.")
                return False

            # 지역사 지분 합계 계산 (결과는 0.49와 같은 소수)
            region_share_sum_decimal = sum(
                comp.get('share', 0) for comp in companies_data
                if region_limit in comp.get('data', {}).get('지역', '')
            )

            # ▼▼▼▼▼ [핵심 수정] ▼▼▼▼▼
            # 비교 및 표시를 위해 소수점 합계를 퍼센트(%)로 변환
            region_share_sum_percent = region_share_sum_decimal * 100.0
            # ▲▲▲▲▲ [핵심 수정] ▲▲▲▲▲

            # 의무비율 미달 시 (49.0 < 49.0 -> False)
            if region_share_sum_percent < duty_ratio:
                reply = QMessageBox.question(self, "의무 비율 경고",
                                             # 경고창에 올바른 퍼센트 값을 표시
                                             f"필수 지역 '{region_limit}' 업체의 지분 합계가 {region_share_sum_percent:.2f}%입니다.\n"
                                             f"의무 비율({duty_ratio}%)에 미달합니다.\n\n"
                                             "감점을 감수하고 계속 진행하시겠습니까?",
                                             QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                                             QMessageBox.StandardButton.No)

                if reply == QMessageBox.StandardButton.No:
                    return False

        return True

    def generate_excel_report(self):
        """사용자가 제공한 최종 보고서 양식(시공실적 포함)에 맞춰 엑셀 파일을 생성합니다."""
        if not self.result_widgets:
            QMessageBox.warning(self, "알림", "먼저 '결과 표 추가' 버튼으로 내보낼 결과를 추가해주세요.")
            return

        # 1. 파일 저장 경로 설정
        safe_title = "".join(c for c in self.gongo_title_entry.text() if c not in r'<>:"/\|?*')
        default_filename = f"{safe_title}.xlsx"
        save_path, _ = QFileDialog.getSaveFileName(self, "엑셀 보고서 저장", default_filename, "Excel Files (*.xlsx)")
        if not save_path:
            return

        try:
            # 2. 템플릿 파일 불러오기
            template_path = resource_path("haeng_template.xlsx")
            wb = load_workbook(template_path)
            ws = wb.active

            # 3. 상단 고정 정보 채우기
            ws['D2'] = utils.parse_amount(self.estimation_price_entry.text())
            ws['M1'] = f"{self.gongo_no_entry.text()} {self.gongo_title_entry.text()}"
            if self.bid_opening_date and self.bid_opening_date.isValid():
                ws['P2'] = self.bid_opening_date.toString("yyyy-MM-dd HH:mm")

            # 4. 데이터 채우기
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            region_limit = self.region_limit_combo.currentText()
            wrap_alignment = Alignment(vertical='center', wrap_text=True)


            # 목록에 있는 모든 컨소시엄 결과에 대해 반복 (5행부터 시작)
            for index, result_widget in enumerate(self.result_widgets):
                current_row = 5 + index
                result_data = result_widget.result_data
                details = result_data.get("company_details", [])

                # 업체별 상세정보 기록
                for comp_detail in details:
                    role = comp_detail.get('role')

                    # 1. 업체명에서 법인 형태 제거 (기존과 동일)
                    original_name = comp_detail.get('name', '')
                    company_name = re.sub(r'\s*㈜\s*|\s*\((주|유|합|재)\)\s*|\s*(주|유|합|재)식회사\s*', '', original_name).strip()

                    # ▼▼▼▼▼ [핵심 추가] 비고란에서 담당자 이름 추출 ▼▼▼▼▼
                    remarks = comp_detail.get('data', {}).get('비고', '')
                    manager_name = None
                    if remarks:
                        # '김OO', '김OO팀장' 등 2~4글자의 한글 이름을 찾는 정규표현식
                        match = re.search(r'([가-힣]{2,4})(님|팀장|실장|부장|과장|대리|주임|사원)?', remarks)
                        if match:
                            manager_name = match.group(1) # '김장섭' 부분만 추출

                        # [디버깅용 코드 추가]
                        print(f"회사: {company_name}, 비고: '{remarks}', 추출된 담당자: {manager_name}")

                    # 최종적으로 셀에 들어갈 텍스트 조합
                    final_cell_text = company_name
                    if manager_name:
                        final_cell_text += f"\n{manager_name}" # 줄바꿈 문자로 이름 추가

                    company_region = comp_detail.get('data', {}).get('지역', '')

                    if role == "대표사":
                        cell = ws.cell(current_row, 3, value=final_cell_text)
                        cell.alignment = wrap_alignment# C열
                    elif role.startswith("구성사"):
                        try:
                            col_offset = 3 + int(role.split(' ')[1])
                            cell = ws.cell(current_row, col_offset, value=final_cell_text)
                            cell.alignment = wrap_alignment
                        except:
                            continue

                    if region_limit != "전체" and region_limit in company_region:
                        cell.fill = yellow_fill

                    # I,J,K... : 지분율
                    share = comp_detail.get('share', 0)

                    # ▼▼▼▼▼ [디버깅] 엑셀에 쓰기 직전의 값을 확인합니다 ▼▼▼▼▼
                    print(f"[디버깅] 엑셀에 쓸 지분율 값: {share} (타입: {type(share)})")
                    # ▲▲▲▲▲ 여기까지 추가 ▲▲▲▲▲

                    if role == "대표사":
                        # [수정] 숫자 값을 그대로 셀에 쓰고, 셀 서식은 '백분율'로 지정
                        ws.cell(current_row, 9, value=share).number_format = '0.00%'
                    elif role.startswith("구성사"):
                        try:
                            col_offset = 9 + int(role.split(' ')[1])
                            ws.cell(current_row, col_offset, value=share).number_format = '0.00%'
                        except:
                            continue

                    # P,Q,R... : 경영상태 점수
                    biz_details = comp_detail.get('business_score_details', {})
                    biz_score = biz_details.get('total', 0)
                    if role == "대표사":
                        ws.cell(current_row, 16, value=biz_score)  # P열
                    elif role.startswith("구성사"):
                        try:
                            col_offset = 16 + int(role.split(' ')[1])
                            ws.cell(current_row, col_offset, value=biz_score)
                        except:
                            continue

                    # ▼▼▼▼▼ [추가] W,X,Y... : 5년 실적 ▼▼▼▼▼
                    performance_5y = comp_detail.get('performance_5y', 0)
                    if role == "대표사":
                        ws.cell(current_row, 23, value=performance_5y).number_format = '#,##0'  # W열
                    elif role.startswith("구성사"):
                        try:
                            col_offset = 23 + int(role.split(' ')[1])  # X, Y, Z...열
                            ws.cell(current_row, col_offset, value=performance_5y).number_format = '#,##0'
                        except:
                            continue
                    # ▲▲▲▲▲ [추가] 여기까지 ▲▲▲▲▲

            # 5. 파일 저장
            wb.save(save_path)
            QMessageBox.information(self, "성공", f"엑셀 보고서가 성공적으로 저장되었습니다.\n경로: {save_path}")

        except FileNotFoundError:
            QMessageBox.critical(self, "템플릿 파일 오류",
                                 f"템플릿 파일('haeng_template.xlsx')을 찾을 수 없습니다.\n프로젝트 폴더에 파일이 있는지 확인해주세요.")
        except Exception as e:
            QMessageBox.critical(self, "오류", f"파일 저장 중 오류가 발생했습니다: {e}")



    def open_result_management_dialog(self):
        """결과 관리 새 창을 엽니다."""
        dialog = ResultManagementDialog(self.result_widgets, self)
        # 새 창에서 데이터가 변경되면, on_results_updated 함수를 실행하도록 연결
        dialog.results_updated.connect(self.on_results_updated)
        dialog.exec()

    def on_results_updated(self, updated_widgets):
        """새 창에서 변경된 결과 목록을 현재 목록에 반영합니다."""
        self.result_widgets = updated_widgets
        QMessageBox.information(self, "반영 완료", f"변경된 결과({len(self.result_widgets)}건)가 반영되었습니다.")

    def update_summary_display(self):
        """하단의 요약 정보 텍스트 상자를 현재 결과 목록에 맞게 업데이트합니다."""
        if not self.result_widgets:
            self.summary_display.clear()
            # Placeholder 텍스트를 다시 보여주기 위해 setPlaceholderText를 사용할 수 있습니다.
            self.summary_display.setPlaceholderText("여기에 추가된 협정 결과가 요약되어 표시됩니다.")
            return

        summary_lines = []
        for i, widget in enumerate(self.result_widgets):
            data = widget.result_data
            details = data.get("company_details", [])

            company_parts = []
            for comp in details:
                name = comp.get('name', '')
                share_percent = comp.get('share', 0) * 100.0
                # '%g' 포맷을 사용하여 불필요한 .0을 제거 (예: 49.0% -> 49%)
                company_parts.append(f"{name}({'%g' % share_percent}%)")

            line = f"<b>[협정 {i + 1}]</b> " + ", ".join(company_parts)
            summary_lines.append(line)

        self.summary_display.setHtml("<br>".join(summary_lines))

    def load_and_recalculate_consortium(self):
        """저장된 협정 파일을 불러와 현재 공고 기준으로 재계산하고 목록에 추가합니다."""
        # ▼▼▼▼▼ [핵심] 덮어쓰기/추가하기 확인 로직 ▼▼▼▼▼
        if self.result_widgets:
            reply = QMessageBox.question(self, "불러오기 방식 선택",
                                         "현재 목록에 추가된 협정이 있습니다.\n\n"
                                         " - [Yes]를 누르면 현재 목록을 지우고 새로 덮어씁니다.\n"
                                         " - [No]를 누르면 현재 목록 뒤에 이어서 추가합니다.\n"
                                         " - [Cancel]을 누르면 작업을 취소합니다.",
                                         QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No | QMessageBox.StandardButton.Cancel,
                                         QMessageBox.StandardButton.Cancel)

            if reply == QMessageBox.StandardButton.Yes:
                self.result_widgets.clear()  # 덮어쓰기 (기존 목록 삭제)
            elif reply == QMessageBox.StandardButton.No:
                pass  # 추가하기 (아무것도 하지 않음)
            else:
                return  # 취소
        # ▲▲▲▲▲ [핵심] 여기까지 ▲▲▲▲▲

        popup = LoadConsortiumPopup(self.mode, self)
        if not popup.exec():
            return

        selected_data_list = popup.get_selected_data()
        if not selected_data_list:
            return

        # --- 공고 정보 수집 (기존과 동일) ---
        if not self.announcement_date_modified: QMessageBox.warning(self, "입력 필요",
                                                                    "정확한 계산을 위해 '공고일'을 반드시 설정해주세요."); return
        announcement_date = self.announcement_date_edit.date().toPython()
        selected_rule_key = self.rule_combo.currentText()
        if not selected_rule_key: QMessageBox.warning(self, "선택 오류", "심사 기준을 선택하세요."); return
        rule_info = (self.mode, selected_rule_key)
        estimation_price_val = utils.parse_amount(self.estimation_price_entry.text())
        if not estimation_price_val: QMessageBox.warning(self, "입력 오류", "추정가격을 정확히 입력해주세요."); return
        base_amount_val = utils.parse_amount(self.base_amount_entry.text())
        tuchal_amount_text = self.tuchal_amount_label.text().replace("<b>", "").replace("</b>", "").replace(" 원",
                                                                                                            "").replace(
            ",", "")
        tuchal_amount_val = utils.parse_amount(tuchal_amount_text) or 0
        price_data = {"estimation_price": estimation_price_val, "notice_base_amount": base_amount_val,
                      "tuchal_amount": tuchal_amount_val}
        sipyung_info = {"is_limited": self.sipyung_limit_check.isChecked(),
                        "limit_amount": utils.parse_amount(self.sipyung_limit_amount.text()) or 0,
                        "method": "비율제" if self.ratio_method_radio.isChecked() else "합산제",
                        "tuchal_amount": price_data["tuchal_amount"]}
        region_limit = self.region_limit_combo.currentText()

        # --- 협정 재구성 및 재계산 ---
        newly_added_count = 0
        for saved_session_data in selected_data_list:
            loaded_consortiums_info = saved_session_data.get("consortiums", [])

            for result_data in loaded_consortiums_info:
                reconstructed_companies = result_data.get("company_details", [])

                if not reconstructed_companies: continue

                new_result = calculation_logic.calculate_consortium(reconstructed_companies, price_data,
                                                                    announcement_date, rule_info, sipyung_info,
                                                                    region_limit)
                if not new_result:
                    QMessageBox.warning(self, "계산 실패", "불러온 협정을 재계산하는 중 오류가 발생했습니다.")
                    continue

                new_result['gongo_title'] = self.gongo_title_entry.text()
                new_result['gongo_no'] = self.gongo_no_entry.text()

                widget = QFrame()
                widget.result_data = new_result
                self.result_widgets.append(widget)
                newly_added_count += 1

        if newly_added_count > 0:
            self.update_summary_display()
            QMessageBox.information(self, "불러오기 완료",
                                    f"선택한 {len(selected_data_list)}개 파일에서 총 {newly_added_count}개의 컨소시엄을 목록에 추가했습니다.")